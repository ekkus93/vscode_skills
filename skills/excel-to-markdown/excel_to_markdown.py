import argparse
import datetime as dt
import os
import re
import sys
from itertools import zip_longest
from typing import Any


def default_output_path(input_file: str) -> str:
    base, _ = os.path.splitext(input_file)
    return base + ".md"


def title_from_path(input_file: str) -> str:
    stem = os.path.splitext(os.path.basename(input_file))[0]
    return re.sub(r"[-_]+", " ", stem).strip() or "Workbook"


def stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def trim_rows(rows: list[list[str]]) -> list[list[str]]:
    trimmed_rows = [list(row) for row in rows]
    while trimmed_rows and not any(cell.strip() for cell in trimmed_rows[-1]):
        trimmed_rows.pop()
    if not trimmed_rows:
        return []
    max_columns = 0
    for row in trimmed_rows:
        for index in range(len(row) - 1, -1, -1):
            if row[index].strip():
                max_columns = max(max_columns, index + 1)
                break
    if max_columns == 0:
        return []
    return [row[:max_columns] + [""] * max(0, max_columns - len(row)) for row in trimmed_rows]


def escape_markdown_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")


def normalize_header_row(row: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, cell in enumerate(row, start=1):
        header = cell.strip() or f"Column {index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            header = f"{header} ({count + 1})"
        headers.append(header)
    return headers


def sheet_to_markdown(rows: list[list[str]]) -> str:
    cleaned_rows = trim_rows(rows)
    if not cleaned_rows:
        return "_Empty sheet._"

    first_non_empty = next(
        (index for index, row in enumerate(cleaned_rows) if any(cell.strip() for cell in row)),
        None,
    )
    if first_non_empty is None:
        return "_Empty sheet._"

    content_rows = cleaned_rows[first_non_empty:]
    column_count = max(len(row) for row in content_rows)
    padded_rows = [row + [""] * (column_count - len(row)) for row in content_rows]

    if len(padded_rows) >= 2:
        header = normalize_header_row(padded_rows[0])
        data_rows = padded_rows[1:]
    else:
        header = [f"Column {index}" for index in range(1, column_count + 1)]
        data_rows = padded_rows

    lines = [
        "| " + " | ".join(escape_markdown_cell(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in data_rows:
        lines.append("| " + " | ".join(escape_markdown_cell(cell) for cell in row) + " |")
    return "\n".join(lines)


def workbook_to_markdown(title: str, sheets: list[tuple[str, list[list[str]]]]) -> str:
    parts = [f"# {title}", ""]
    for index, (sheet_name, rows) in enumerate(sheets):
        if index:
            parts.append("")
        parts.append(f"## {sheet_name}")
        parts.append("")
        parts.append(sheet_to_markdown(rows))
    return "\n".join(parts).rstrip() + "\n"


def load_xlsx_sheets(input_file: str) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            "Missing required Python package: openpyxl. "
            "Install it with python3 -m pip install openpyxl"
        ) from exc

    data_workbook = load_workbook(filename=input_file, read_only=True, data_only=True)
    formula_workbook = load_workbook(filename=input_file, read_only=True, data_only=False)
    try:
        sheets: list[tuple[str, list[list[str]]]] = []
        for data_sheet, formula_sheet in zip(
            data_workbook.worksheets, formula_workbook.worksheets, strict=False
        ):
            rows: list[list[str]] = []
            for data_row, formula_row in zip_longest(
                data_sheet.iter_rows(values_only=True),
                formula_sheet.iter_rows(values_only=True),
                fillvalue=(),
            ):
                row_values: list[str] = []
                width = max(len(data_row), len(formula_row))
                for index in range(width):
                    data_value = data_row[index] if index < len(data_row) else None
                    formula_value = formula_row[index] if index < len(formula_row) else None
                    value = data_value
                    if (
                        value is None
                        and isinstance(formula_value, str)
                        and formula_value.startswith("=")
                    ):
                        value = formula_value
                    row_values.append(stringify_value(value))
                rows.append(row_values)
            sheets.append((data_sheet.title, trim_rows(rows)))
        return sheets
    finally:
        data_workbook.close()
        formula_workbook.close()


def load_xls_sheets(input_file: str) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            "Missing required Python package: xlrd. Install it with python3 -m pip install xlrd"
        ) from exc

    workbook = xlrd.open_workbook(input_file)
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet in workbook.sheets():
        rows: list[list[str]] = []
        for row_index in range(sheet.nrows):
            row_values: list[str] = []
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row_values.append(
                        stringify_value(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                    )
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row_values.append("TRUE" if cell.value else "FALSE")
                else:
                    row_values.append(stringify_value(cell.value))
            rows.append(row_values)
        sheets.append((sheet.name, trim_rows(rows)))
    return sheets


def read_workbook_sheets(input_file: str) -> list[tuple[str, list[list[str]]]]:
    lower_name = input_file.lower()
    if lower_name.endswith(".xlsx"):
        return load_xlsx_sheets(input_file)
    if lower_name.endswith(".xls"):
        return load_xls_sheets(input_file)
    raise RuntimeError("Input file must end in .xls or .xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert Excel workbooks to Markdown")
    parser.add_argument("input_file")
    parser.add_argument("output_file", nargs="?")
    args = parser.parse_args()

    input_file = os.path.abspath(args.input_file)
    output_file = (
        os.path.abspath(args.output_file)
        if args.output_file
        else default_output_path(input_file)
    )

    if not os.path.isfile(input_file):
        print(f"Input file not found: {input_file}", file=sys.stderr)
        return 2

    if not input_file.lower().endswith((".xls", ".xlsx")):
        print("Input file must end in .xls or .xlsx", file=sys.stderr)
        return 2

    try:
        sheets = read_workbook_sheets(input_file)
        markdown = workbook_to_markdown(title_from_path(input_file), sheets)
        with open(output_file, "w", encoding="utf-8") as handle:
            handle.write(markdown)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    if not os.path.isfile(output_file) or os.path.getsize(output_file) == 0:
        print("Markdown output was not created", file=sys.stderr)
        return 1

    print(output_file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())